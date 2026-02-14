from flask import Flask, render_template_string, request, redirect, url_for, session, flash, send_file
import json
import os
from datetime import datetime, timedelta
import csv
from io import BytesIO
from functools import wraps
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

app = Flask(__name__)
app.secret_key = 'your-secret-key-change-this-in-production'

# Session configuration
app.permanent_session_lifetime = timedelta(minutes=30)  # adjust timeout as needed
app.config['SESSION_COOKIE_HTTPONLY'] = True

# Helper: require admin login
def login_required(view_func):
    @wraps(view_func)
    def wrapped_view(*args, **kwargs):
        if 'admin' not in session:
            return redirect(url_for('admin_login'))
        return view_func(*args, **kwargs)
    return wrapped_view


# Helper: clean up admins with no questions added in 24+ hours
def cleanup_inactive_admins():
    admins = load_admins()
    questions = load_questions()
    now = datetime.utcnow()
    cutoff = now - timedelta(hours=24)
    
    to_delete = []
    for admin_id, admin_data in admins.items():
        # Get admin's question count
        admin_questions = questions.get(admin_id, [])
        if not admin_questions:  # No questions
            created_at_str = admin_data.get('created_at', '')
            try:
                created_at = datetime.fromisoformat(created_at_str)
                if created_at < cutoff:  # Created more than 24 hours ago
                    to_delete.append(admin_id)
            except Exception:
                # If parsing fails, skip
                pass
    
    # Delete inactive admins and their data
    for admin_id in to_delete:
        del admins[admin_id]
        questions.pop(admin_id, None)
        allowed = load_allowed()
        allowed.pop(admin_id, None)
        save_allowed(allowed)
        quiz_settings = load_quiz_settings()
        quiz_settings.pop(admin_id, None)
        save_quiz_settings(quiz_settings)
        all_answers = load_answers()
        all_answers.pop(admin_id, None)
        save_answers(all_answers)
    
    if to_delete:
        save_admins(admins)
        save_questions(questions)


# Helper: send welcome email to admin
def send_admin_welcome_email(admin_username, admin_email):
    # Email configuration - read from environment or use defaults
    sender_email = os.getenv('SMTP_EMAIL', 'your_email@gmail.com')
    sender_password = os.getenv('SMTP_PASSWORD', '')
    smtp_server = os.getenv('SMTP_SERVER', 'smtp.gmail.com')
    smtp_port = int(os.getenv('SMTP_PORT', '587'))
    
    # If no credentials configured, just log (don't fail)
    if not sender_password or sender_email == 'your_email@gmail.com':
        print(f"[EMAIL - NOT SENT] Email credentials not configured.")
        print(f"To: {admin_email}")
        print(f"Subject: Welcome to Quiz Management System")
        print(f"Username: {admin_username}")
        return
    
    subject = "Welcome to Quiz Management System"
    body = f"""
Dear Administrator,

Your account has been successfully created!

Username: {admin_username}
Email: {admin_email}

You can now log in to the admin panel at: http://127.0.0.1:5000/admin/login

Please keep your credentials safe and secure.

Best regards,
Quiz Management System Team
"""
    
    try:
        # Create email message
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = admin_email
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain'))
        
        # Send email
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.send_message(msg)
        print(f"✓ Welcome email sent to {admin_email}")
    except Exception as e:
        print(f"✗ Error sending email to {admin_email}: {e}")


@app.before_request
def check_session_timeout():
    # If admin is logged in, enforce inactivity timeout and refresh last_active
    if 'admin' in session:
        last = session.get('last_active')
        if last:
            try:
                last_dt = datetime.fromisoformat(last)
                if datetime.utcnow() - last_dt > app.permanent_session_lifetime:
                    session.pop('admin', None)
                    session.pop('last_active', None)
                    flash('Session timed out. Please log in again.', 'error')
                    return redirect(url_for('admin_login'))
            except Exception:
                # If parsing fails, clear the timestamp and continue
                session.pop('last_active', None)
        # Update last activity timestamp to extend session
        session['last_active'] = datetime.utcnow().isoformat()

# File to store data
ADMINS_FILE = 'admins.json'
QUESTIONS_FILE = 'questions.json'
ANSWERS_FILE = 'user_answers.json'
QUIZ_SETTINGS_FILE = 'quiz_settings.json'
ALLOWED_FILE = 'allowed_students.json'

# Initialize files if they don't exist
def init_files():
    if not os.path.exists(ADMINS_FILE):
        with open(ADMINS_FILE, 'w') as f:
            json.dump({}, f)
    if not os.path.exists(QUESTIONS_FILE):
        with open(QUESTIONS_FILE, 'w') as f:
            json.dump({}, f)
    if not os.path.exists(ANSWERS_FILE):
        with open(ANSWERS_FILE, 'w') as f:
            json.dump({}, f)
    if not os.path.exists(QUIZ_SETTINGS_FILE):
        with open(QUIZ_SETTINGS_FILE, 'w') as f:
            json.dump({}, f)
    if not os.path.exists(ALLOWED_FILE):
        with open(ALLOWED_FILE, 'w') as f:
            json.dump({}, f)
    
    # Migrate old data format to new format
    migrate_old_data()

def migrate_old_data():
    """Convert old list format to new dictionary format"""
    # Migrate questions
    try:
        with open(QUESTIONS_FILE, 'r') as f:
            data = json.load(f)
            if isinstance(data, list):  # Old format
                # Move old questions to a default admin
                new_data = {'default_admin': data}
                with open(QUESTIONS_FILE, 'w') as fw:
                    json.dump(new_data, fw, indent=2)
                
                # Create default admin if questions exist
                if data:
                    admins = load_admins()
                    if 'default_admin' not in admins:
                        admins['default_admin'] = {
                            'password': 'admin123',
                            'created_at': datetime.now().isoformat()
                        }
                        save_admins(admins)
    except:
        pass
    
    # Migrate answers
    try:
        with open(ANSWERS_FILE, 'r') as f:
            data = json.load(f)
            if isinstance(data, list):  # Old format
                # Move old answers to default admin
                new_data = {'default_admin': data}
                with open(ANSWERS_FILE, 'w') as fw:
                    json.dump(new_data, fw, indent=2)
    except:
        pass

# Load admins
def load_admins():
    with open(ADMINS_FILE, 'r') as f:
        return json.load(f)

# Load allowed students
def load_allowed():
    with open(ALLOWED_FILE, 'r') as f:
        return json.load(f)

def save_allowed(data):
    with open(ALLOWED_FILE, 'w') as f:
        json.dump(data, f, indent=2)

# Save admins
def save_admins(admins):
    with open(ADMINS_FILE, 'w') as f:
        json.dump(admins, f, indent=2)

# Load questions
def load_questions():
    with open(QUESTIONS_FILE, 'r') as f:
        return json.load(f)

# Save questions
def save_questions(questions):
    with open(QUESTIONS_FILE, 'w') as f:
        json.dump(questions, f, indent=2)

# Load answers
def load_answers():
    with open(ANSWERS_FILE, 'r') as f:
        return json.load(f)

# Save answers
def save_answers(answers):
    with open(ANSWERS_FILE, 'w') as f:
        json.dump(answers, f, indent=2)

# Load quiz settings
def load_quiz_settings():
    with open(QUIZ_SETTINGS_FILE, 'r') as f:
        return json.load(f)

# Save quiz settings
def save_quiz_settings(settings):
    with open(QUIZ_SETTINGS_FILE, 'w') as f:
        json.dump(settings, f, indent=2)

# HTML Templates
HOME_TEMPLATE = '''
<!DOCTYPE html>
<html>
<head>
    <title>Quiz System</title>
    <link rel="icon" type="image/svg+xml" href="{{ url_for('static', filename='favicon.svg') }}">
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            display: flex;
            align-items: center;
            justify-content: center;
            padding: 20px;
        }
        .home-container {
            background: white;
            padding: 50px 40px;
            border-radius: 20px;
            box-shadow: 0 20px 60px rgba(0, 0, 0, 0.3);
            text-align: center;
            max-width: 600px;
            animation: slideUp 0.6s ease-out;
        }
        @keyframes slideUp {
            from {
                opacity: 0;
                transform: translateY(30px);
            }
            to {
                opacity: 1;
                transform: translateY(0);
            }
        }
        h1 {
            color: #333;
            font-size: 2.5rem;
            font-weight: 700;
            margin-bottom: 15px;
        }
        .subtitle {
            color: #666;
            font-size: 1.1rem;
            margin-bottom: 40px;
        }
        .btn-container {
            display: flex;
            gap: 20px;
            flex-direction: column;
        }
        @media (min-width: 576px) {
            .btn-container {
                flex-direction: row;
                justify-content: center;
            }
        }
        .btn-modern {
            padding: 15px 40px;
            font-size: 1.1rem;
            font-weight: 600;
            border: none;
            border-radius: 12px;
            cursor: pointer;
            text-decoration: none;
            color: white;
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
            box-shadow: 0 4px 15px rgba(0, 0, 0, 0.2);
            display: inline-block;
            flex: 1;
        }
        .btn-modern:hover {
            transform: translateY(-3px);
            box-shadow: 0 8px 20px rgba(0, 0, 0, 0.3);
        }
        .btn-admin {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        }
        .btn-user {
            background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);
        }
        .btn-modern:active {
            transform: translateY(-1px);
        }
    </style>
</head>
<body>
    <div class="home-container">
        <h1>📚 Quiz Management System</h1>
        <p class="subtitle">Welcome! Choose your role to continue</p>
        <div class="btn-container">
            <a href="{{ url_for('admin_login') }}" class="btn-modern btn-admin">🔐 Admin Panel</a>
            <a href="{{ url_for('start') }}" class="btn-modern btn-user">✏️ Take Quiz</a>
        </div>
    </div>
</body>
</html>
'''

START_TEMPLATE = '''
<!DOCTYPE html>
<html>
<head>
    <title>Enter Details</title>
    <link rel="icon" type="image/svg+xml" href="{{ url_for('static', filename='favicon.svg') }}">
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            display: flex;
            align-items: center;
            justify-content: center;
            padding: 20px;
        }
        .start-container {
            background: white;
            padding: 40px;
            border-radius: 20px;
            box-shadow: 0 20px 60px rgba(0, 0, 0, 0.3);
            max-width: 500px;
            width: 100%;
            animation: slideUp 0.6s ease-out;
        }
        @keyframes slideUp {
            from {
                opacity: 0;
                transform: translateY(30px);
            }
            to {
                opacity: 1;
                transform: translateY(0);
            }
        }
        h2 {
            color: #333;
            font-weight: 700;
            margin-bottom: 30px;
            text-align: center;
        }
        .form-group {
            margin-bottom: 20px;
        }
        .form-control {
            padding: 12px 16px;
            border: 2px solid #e0e0e0;
            border-radius: 10px;
            font-size: 1rem;
            transition: all 0.3s ease;
        }
        .form-control:focus {
            border-color: #667eea;
            box-shadow: 0 0 0 3px rgba(102, 126, 234, 0.1);
        }
        .alert {
            border-radius: 10px;
            border: none;
            margin-bottom: 20px;
            animation: slideDown 0.3s ease;
        }
        @keyframes slideDown {
            from {
                opacity: 0;
                transform: translateY(-10px);
            }
            to {
                opacity: 1;
                transform: translateY(0);
            }
        }
        .alert-danger {
            background: #ffe5e5;
            color: #c33;
        }
        .alert-success {
            background: #e5f8e5;
            color: #2d7a2d;
        }
        .button-group {
            display: flex;
            gap: 12px;
            margin-top: 30px;
        }
        .btn-continue {
            flex: 1;
            padding: 12px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            border: none;
            border-radius: 10px;
            font-weight: 600;
            cursor: pointer;
            transition: all 0.3s ease;
        }
        .btn-continue:hover:not(:disabled) {
            transform: translateY(-2px);
            box-shadow: 0 8px 20px rgba(102, 126, 234, 0.4);
        }
        .btn-continue:disabled {
            opacity: 0.6;
            cursor: not-allowed;
        }
        .btn-cancel {
            flex: 1;
            padding: 12px;
            background: #f0f0f0;
            color: #333;
            border: 2px solid #e0e0e0;
            border-radius: 10px;
            font-weight: 600;
            cursor: pointer;
            text-decoration: none;
            display: flex;
            align-items: center;
            justify-content: center;
            transition: all 0.3s ease;
        }
        .btn-cancel:hover {
            background: #e8e8e8;
            transform: translateY(-2px);
        }
    </style>
</head>
<body>
    <div class="start-container">
        <h2>👤 Enter Your Details</h2>
        {% with messages = get_flashed_messages(with_categories=true) %}
            {% if messages %}
                {% for cat, m in messages %}
                    <div class="alert alert-{{ 'danger' if cat=='error' else 'success' }}">{{ m }}</div>
                {% endfor %}
            {% endif %}
        {% endwith %}
        <form method="POST" action="{{ url_for('start') }}">
            <div class="form-group">
                <input type="text" name="student_name" class="form-control" placeholder="Full Name" required autocomplete="off">
            </div>
            <div class="form-group">
                <input type="text" name="student_id" class="form-control" placeholder="Registration / Reg No" required autocomplete="off">
            </div>
            <div class="button-group">
                <button type="submit" class="btn-continue">Continue →</button>
                <a href="{{ url_for('home') }}" class="btn-cancel">Cancel</a>
            </div>
        </form>
    </div>
</body>
</html>
'''

ADMIN_LOGIN_TEMPLATE = '''
<!DOCTYPE html>
<html>
<head>
    <title>Admin Access</title>
    <link rel="icon" type="image/svg+xml" href="{{ url_for('static', filename='favicon.svg') }}">
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            display: flex;
            align-items: center;
            justify-content: center;
            padding: 20px;
        }
        .login-container {
            background: white;
            padding: 0;
            border-radius: 20px;
            box-shadow: 0 20px 60px rgba(0, 0, 0, 0.3);
            max-width: 500px;
            width: 100%;
            overflow: hidden;
            animation: slideUp 0.6s ease-out;
        }
        @keyframes slideUp {
            from {
                opacity: 0;
                transform: translateY(30px);
            }
            to {
                opacity: 1;
                transform: translateY(0);
            }
        }
        .tab-header {
            display: flex;
            background: #f8f9fa;
            border-bottom: 2px solid #e0e0e0;
        }
        .tab-btn {
            flex: 1;
            padding: 16px;
            background: transparent;
            border: none;
            cursor: pointer;
            font-weight: 600;
            color: #666;
            font-size: 1rem;
            transition: all 0.3s ease;
            position: relative;
        }
        .tab-btn.active {
            color: #667eea;
        }
        .tab-btn.active::after {
            content: '';
            position: absolute;
            bottom: -2px;
            left: 0;
            right: 0;
            height: 3px;
            background: #667eea;
        }
        .tab-content {
            padding: 40px;
            display: none;
            animation: fadeIn 0.3s ease;
        }
        .tab-content.active {
            display: block;
        }
        @keyframes fadeIn {
            from {
                opacity: 0;
            }
            to {
                opacity: 1;
            }
        }
        h3 {
            color: #333;
            margin-bottom: 30px;
            font-weight: 700;
        }
        .form-group {
            margin-bottom: 18px;
        }
        .form-control {
            padding: 12px 16px;
            border: 2px solid #e0e0e0;
            border-radius: 10px;
            font-size: 1rem;
            transition: all 0.3s ease;
        }
        .form-control:focus {
            border-color: #667eea;
            box-shadow: 0 0 0 3px rgba(102, 126, 234, 0.1);
            outline: none;
        }
        .btn-submit {
            width: 100%;
            padding: 12px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            border: none;
            border-radius: 10px;
            font-weight: 600;
            cursor: pointer;
            font-size: 1rem;
            transition: all 0.3s ease;
            margin-top: 10px;
        }
        .btn-submit:hover {
            transform: translateY(-2px);
            box-shadow: 0 8px 20px rgba(102, 126, 234, 0.4);
        }
        .alert {
            border-radius: 10px;
            border: none;
            margin-bottom: 20px;
            animation: slideDown 0.3s ease;
        }
        @keyframes slideDown {
            from {
                opacity: 0;
                transform: translateY(-10px);
            }
            to {
                opacity: 1;
                transform: translateY(0);
            }
        }
        .alert-danger {
            background: #ffe5e5;
            color: #c33;
        }
        .alert-success {
            background: #e5f8e5;
            color: #2d7a2d;
        }
        .back-link {
            display: block;
            text-align: center;
            margin-top: 20px;
            color: #667eea;
            text-decoration: none;
            font-weight: 600;
            transition: all 0.3s ease;
        }
        .back-link:hover {
            color: #764ba2;
        }
    </style>
    <script>
        function showTab(tabName) {
            document.querySelectorAll('.tab-content').forEach(tab => tab.classList.remove('active'));
            document.querySelectorAll('.tab-btn').forEach(btn => btn.classList.remove('active'));
            document.getElementById(tabName).classList.add('active');
            event.target.classList.add('active');
        }
    </script>
</head>
<body>
    <div class="login-container">
        <div class="tab-header">
            <button class="tab-btn active" onclick="showTab('login-tab')">🔐 Login</button>
            <button class="tab-btn" onclick="showTab('register-tab')">📝 Register</button>
        </div>
        
        {% with messages = get_flashed_messages(with_categories=true) %}
            {% if messages %}
                <div style="padding: 20px 40px; padding-bottom: 0;">
                    {% for category, message in messages %}
                        <div class="alert alert-{{ category }}">{{ message }}</div>
                    {% endfor %}
                </div>
            {% endif %}
        {% endwith %}
        
        <!-- Login Tab -->
        <div id="login-tab" class="tab-content active">
            <h3>Welcome Back</h3>
            <form method="POST" action="{{ url_for('admin_login') }}">
                <input type="hidden" name="action" value="login">
                <div class="form-group">
                    <input type="text" name="username" class="form-control" placeholder="Username" required>
                </div>
                <div class="form-group">
                    <input type="password" name="password" class="form-control" placeholder="Password" required>
                </div>
                <button type="submit" class="btn-submit">Sign In</button>
            </form>
        </div>
        
        <!-- Register Tab -->
        <div id="register-tab" class="tab-content">
            <h3>Create Admin Account</h3>
            <form method="POST" action="{{ url_for('admin_login') }}">
                <input type="hidden" name="action" value="register">
                <div class="form-group">
                    <input type="text" name="username" class="form-control" placeholder="Username" required>
                </div>
                <div class="form-group">
                    <input type="email" name="email" class="form-control" placeholder="Email Address" required>
                </div>
                <div class="form-group">
                    <input type="tel" name="phone" class="form-control" placeholder="Phone Number" required>
                </div>
                <div class="form-group">
                    <input type="password" name="password" class="form-control" placeholder="Password" required>
                </div>
                <div class="form-group">
                    <input type="password" name="confirm_password" class="form-control" placeholder="Confirm Password" required>
                </div>
                <button type="submit" class="btn-submit" style="background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);">Create Account</button>
            </form>
        </div>
        
        <div style="padding: 0 40px 20px;">
            <a href="{{ url_for('home') }}" class="back-link">← Back to Home</a>
        </div>
    </div>
</body>
</html>
'''

ADMIN_PANEL_TEMPLATE = '''
<!DOCTYPE html>
<html>
<head>
    <title>Admin Panel</title>
    <link rel="icon" type="image/svg+xml" href="{{ url_for('static', filename='favicon.svg') }}">
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 30px 20px;
        }
        .admin-wrapper {
            max-width: 1400px;
            margin: 0 auto;
        }
        .header-bar {
            background: white;
            padding: 25px;
            border-radius: 15px;
            box-shadow: 0 10px 30px rgba(0, 0, 0, 0.2);
            margin-bottom: 30px;
            display: flex;
            justify-content: space-between;
            align-items: center;
            animation: slideDown 0.5s ease;
        }
        @keyframes slideDown {
            from { opacity: 0; transform: translateY(-20px); }
            to { opacity: 1; transform: translateY(0); }
        }
        .header-title h1 {
            color: #333;
            font-weight: 700;
            margin: 0;
            font-size: 2rem;
        }
        .header-title p {
            color: #666;
            margin: 5px 0 0 0;
            font-size: 0.95rem;
        }
        .logout-btn {
            background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);
            color: white;
            border: none;
            padding: 12px 30px;
            border-radius: 10px;
            font-weight: 600;
            cursor: pointer;
            text-decoration: none;
            transition: all 0.3s ease;
            display: inline-block;
        }
        .logout-btn:hover {
            transform: translateY(-2px);
            box-shadow: 0 8px 20px rgba(245, 87, 108, 0.4);
        }
        .status-card {
            background: white;
            padding: 20px;
            border-radius: 12px;
            box-shadow: 0 5px 15px rgba(0, 0, 0, 0.1);
            margin-bottom: 20px;
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            animation: slideUp 0.5s ease 0.1s both;
        }
        @keyframes slideUp {
            from { opacity: 0; transform: translateY(20px); }
            to { opacity: 1; transform: translateY(0); }
        }
        .stat-item {
            text-align: center;
            padding: 20px;
            border-radius: 10px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
        }
        .stat-label {
            font-size: 0.9rem;
            opacity: 0.9;
            margin-bottom: 10px;
        }
        .stat-value {
            font-size: 2rem;
            font-weight: 700;
        }
        .tab-navigation {
            display: flex;
            gap: 10px;
            margin-bottom: 25px;
            background: white;
            padding: 15px;
            border-radius: 12px;
            box-shadow: 0 5px 15px rgba(0, 0, 0, 0.1);
            flex-wrap: wrap;
        }
        .tab-btn {
            padding: 12px 24px;
            background: #f0f0f0;
            border: 2px solid transparent;
            border-radius: 8px;
            cursor: pointer;
            font-weight: 600;
            color: #666;
            transition: all 0.3s ease;
        }
        .tab-btn:hover {
            background: #e8e8e8;
            border-color: #667eea;
        }
        .tab-btn.active {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            border-color: #667eea;
        }
        .content-panel {
            background: white;
            padding: 30px;
            border-radius: 15px;
            box-shadow: 0 10px 30px rgba(0, 0, 0, 0.2);
            display: none;
            animation: fadeIn 0.4s ease;
        }
        .content-panel.active {
            display: block;
        }
        @keyframes fadeIn {
            from { opacity: 0; }
            to { opacity: 1; }
        }
        .container {
            background: transparent;
            padding: 0;
            border-radius: 0;
            box-shadow: none;
            margin-bottom: 0;
        }
        h2 {
            color: #333;
            border-bottom: none;
            padding-bottom: 0;
            margin-bottom: 20px;
        }
        .admin-info {
            background: #f0f0f0;
            padding: 10px 15px;
            border-radius: 5px;
            margin-bottom: 20px;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }
        .form-group {
            margin-bottom: 20px;
        }
        label {
            display: block;
            font-weight: 600;
            color: #333;
            margin-bottom: 8px;
        }
        input, textarea, select {
            width: 100%;
            padding: 12px 16px;
            border: 2px solid #e0e0e0;
            border-radius: 8px;
            font-size: 1rem;
            transition: all 0.3s ease;
            font-family: inherit;
        }
        textarea {
            min-height: 120px;
            resize: vertical;
        }
        input:focus, textarea:focus, select:focus {
            border-color: #667eea;
            box-shadow: 0 0 0 3px rgba(102, 126, 234, 0.1);
            outline: none;
        }
        .btn {
            padding: 12px 30px;
            border: none;
            border-radius: 8px;
            font-weight: 600;
            cursor: pointer;
            text-decoration: none;
            display: inline-flex;
            align-items: center;
            justify-content: center;
            gap: 8px;
            transition: all 0.3s ease;
        }
        .btn-primary {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
        }
        .btn-primary:hover {
            transform: translateY(-2px);
            box-shadow: 0 8px 20px rgba(102, 126, 234, 0.4);
        }
        .btn-secondary {
            background: #f0f0f0;
            color: #333;
            border: 2px solid #e0e0e0;
        }
        .btn-secondary:hover {
            background: #e8e8e8;
            border-color: #999;
        }
        .btn-success {
            background: linear-gradient(135deg, #43e97b 0%, #38f9d7 100%);
            color: white;
        }
        .btn-success:hover {
            transform: translateY(-2px);
            box-shadow: 0 8px 20px rgba(67, 233, 123, 0.4);
        }
        .btn-info {
            background: linear-gradient(135deg, #4facfe 0%, #00f2fe 100%);
            color: white;
        }
        .btn-info:hover {
            transform: translateY(-2px);
            box-shadow: 0 8px 20px rgba(79, 172, 254, 0.4);
        }
        .question-card {
            background: #f8f9fa;
            border-left: 4px solid #667eea;
            padding: 20px;
            border-radius: 8px;
            margin-bottom: 15px;
            transition: all 0.3s ease;
        }
        .question-card:hover {
            box-shadow: 0 5px 15px rgba(0, 0, 0, 0.1);
        }
        .question-title {
            font-weight: 700;
            color: #333;
            margin-bottom: 12px;
        }
        .option-item {
            padding: 8px 12px;
            margin: 6px 0;
            background: white;
            border-radius: 6px;
            border-left: 3px solid #ddd;
        }
        .option-item.correct {
            border-left-color: #27ae60;
            color: #27ae60;
            font-weight: 600;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
        }
        th {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 15px;
            text-align: left;
            font-weight: 600;
        }
        td {
            padding: 12px 15px;
            border-bottom: 1px solid #e0e0e0;
        }
        tr:hover {
            background: #f8f9fa;
        }
        .chart-container {
            background: white;
            padding: 20px;
            border-radius: 8px;
            margin-bottom: 20px;
            box-shadow: 0 2px 8px rgba(0, 0, 0, 0.1);
        }
        .alert-box {
            padding: 16px;
            border-radius: 8px;
            margin-bottom: 20px;
            border-left: 4px solid;
        }
        .alert-box.info {
            background: #e3f2fd;
            border-left-color: #2196f3;
            color: #1565c0;
        }
        .alert-box.success {
            background: #e8f5e9;
            border-left-color: #4caf50;
            color: #2e7d32;
        }
        .alert-box.error {
            background: #ffebee;
            border-left-color: #f44336;
            color: #c62828;
        }
    </style>
    <script>
        function showTab(tabName) {
            document.querySelectorAll('.tab-content').forEach(tab => {
                tab.style.display = 'none';
            });
            document.querySelectorAll('.nav-tab').forEach(btn => {
                btn.classList.remove('active');
            });
            document.getElementById(tabName).style.display = 'block';
            event.target.classList.add('active');
        }
    </script>
    <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.js"></script>
    <script>
        function showTab(tabName) {
            document.querySelectorAll('.content-panel').forEach(p => p.classList.remove('active'));
            document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
            document.getElementById(tabName).classList.add('active');
            event.target.classList.add('active');
        }
    </script>
</head>
<body>
    <div class="admin-wrapper">
        <div class="header-bar">
            <div class="header-title">
                <h1>📊 Admin Dashboard</h1>
                <p>Welcome, <strong>{{ current_admin }}</strong></p>
            </div>
            <a href="{{ url_for('logout') }}" class="logout-btn">🚪 Logout</a>
        </div>

        <div class="status-card">
            <div class="stat-item">
                <div class="stat-label">Questions Set</div>
                <div class="stat-value">{{ questions|length }}</div>
            </div>
            <div class="stat-item" style="background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);">
                <div class="stat-label">Student Submissions</div>
                <div class="stat-value">{{ student_results|length }}</div>
            </div>
            <div class="stat-item" style="background: linear-gradient(135deg, #4facfe 0%, #00f2fe 100%);">
                <div class="stat-label">Submitted by</div>
                <div class="stat-value">{{ submitted_names|length }}</div>
            </div>
        </div>

        <div class="tab-navigation">
            <button class="tab-btn active" onclick="showTab('questions-section')">📋 Questions</button>
            <button class="tab-btn" onclick="showTab('settings-section')">⚙️ Settings</button>
            <button class="tab-btn" onclick="showTab('results-section')">📊 Results</button>
            <button class="tab-btn" onclick="showTab('analytics-section')">📈 Analytics</button>
        </div>

        <div id="questions-section" class="content-panel active">
            <h2>📋 Manage Questions</h2>
            <hr style="margin: 20px 0;">
            
            <h4>Add New Question</h4>
            <form method="POST" action="{{ url_for('admin_panel') }}">
                <div class="form-group">
                    <label>Question Text</label>
                    <textarea name="question" class="form-control" placeholder="Enter your question here..." required></textarea>
                </div>
                
                <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 15px;">
                    <div class="form-group">
                        <label>Option A</label>
                        <input type="text" name="option_a" class="form-control" required>
                    </div>
                    <div class="form-group">
                        <label>Option B</label>
                        <input type="text" name="option_b" class="form-control" required>
                    </div>
                    <div class="form-group">
                        <label>Option C</label>
                        <input type="text" name="option_c" class="form-control" required>
                    </div>
                    <div class="form-group">
                        <label>Option D</label>
                        <input type="text" name="option_d" class="form-control" required>
                    </div>
                </div>
                
                <div class="form-group">
                    <label>Correct Answer</label>
                    <select name="correct_answer" class="form-control" required>
                        <option value="">Select correct answer...</option>
                        <option value="A">A</option>
                        <option value="B">B</option>
                        <option value="C">C</option>
                        <option value="D">D</option>
                    </select>
                </div>
                
                <button type="submit" class="btn btn-primary">✅ Add Question</button>
            </form>

            <hr style="margin: 40px 0;">
            <h4>Your Questions ({{ questions|length }} total)</h4>
            
            {% if questions %}
                {% for q in questions %}
                    <div class="question-card">
                        <div class="question-title">Q{{ loop.index }}: {{ q.question }}</div>
                        <div style="margin: 12px 0;">
                            <div class="option-item {% if q.correct_answer == 'A' %}correct{% endif %}">A) {{ q.options.A }}</div>
                            <div class="option-item {% if q.correct_answer == 'B' %}correct{% endif %}">B) {{ q.options.B }}</div>
                            <div class="option-item {% if q.correct_answer == 'C' %}correct{% endif %}">C) {{ q.options.C }}</div>
                            <div class="option-item {% if q.correct_answer == 'D' %}correct{% endif %}">D) {{ q.options.D }}</div>
                        </div>
                        <form method="POST" action="{{ url_for('delete_question', index=loop.index0) }}" style="display: inline;">
                            <button type="submit" class="btn btn-secondary" style="margin-top: 12px;" onclick="return confirm('Delete this question?')">🗑️ Delete</button>
                        </form>
                    </div>
                {% endfor %}
            {% else %}
                <div class="alert-box info">
                    📚 No questions added yet. Create your first question above!
                </div>
            {% endif %}
        </div>

        <div id="settings-section" class="content-panel">
            <h2>⚙️ Quiz Settings</h2>
            <hr style="margin: 20px 0;">
            
            <h4>Timer Configuration</h4>
            <div style="background: linear-gradient(135deg, #fff5e6 0%, #ffe6cc 100%); padding: 20px; border-radius: 8px; margin-bottom: 20px;">
                <div style="font-weight: 600; color: #333; margin-bottom: 8px;">⏱️ Current Setting:</div>
                {% if quiz_time_limit > 0 %}
                    <span style="display: inline-block; background: #ffc107; color: #000; padding: 6px 12px; border-radius: 20px; font-weight: 600;">{{ quiz_time_limit }} minutes</span>
                    <span style="margin-left: 10px; color: #666;">Students have {{ quiz_time_limit }} mins to complete</span>
                {% else %}
                    <span style="display: inline-block; background: #96f6ff; color: #006666; padding: 6px 12px; border-radius: 20px; font-weight: 600;">No Time Limit</span>
                    <span style="margin-left: 10px; color: #666;">Students can take unlimited time</span>
                {% endif %}
            </div>
            
            <form method="POST" action="{{ url_for('update_quiz_settings') }}">
                <div class="form-group">
                    <label>Set Quiz Timer (minutes)</label>
                    <input type="number" name="time_limit" class="form-control" min="0" max="180" value="{{ quiz_time_limit }}" placeholder="0 = no limit">
                    <small style="color: #666; margin-top: 6px; display: block;">Enter 0 for no limit, or specify minutes (e.g., 30)</small>
                </div>
                <button type="submit" class="btn btn-primary">💾 Save Settings</button>
            </form>

            <hr style="margin: 30px 0;">
            <h4>📥 Upload Student List (CSV)</h4>
            <div style="background: #f0f4ff; padding: 15px; border-radius: 8px; margin-bottom: 20px; border-left: 4px solid #667eea;">
                <strong style="color: #667eea;">📋 CSV Format Instructions:</strong>
                <ul style="color: #555; margin: 10px 0; padding-left: 20px;">
                    <li><strong>First column:</strong> Student Name</li>
                    <li><strong>Second column:</strong> Student ID</li>
                    <li><strong>Separator:</strong> Comma (,) or Tab</li>
                    <li><strong>Example:</strong> <code style="background: white; padding: 2px 6px; border-radius: 3px;">Alfayo3 Nkinda,EASTC/BDTS/24/01034</code></li>
                </ul>
                <strong style="color: #e74c3c;">⚠️ From Excel:</strong> Copy from Excel and paste into .csv file (Excel exports with tabs by default)
            </div>
            
            <form method="POST" action="{{ url_for('upload_students') }}" enctype="multipart/form-data">
                <div class="form-group">
                    <label>Select CSV File</label>
                    <input type="file" name="students_file" class="form-control" accept=".csv,.txt" required>
                </div>
                <button type="submit" class="btn btn-primary">📤 Upload & Replace Student List</button>
            </form>

            {% if allowed_list %}
                <div style="margin-top: 35px; background: #f8f9fa; padding: 20px; border-radius: 8px; border: 2px solid #e0e0e0;">
                    <h4 style="margin-top: 0; color: #333;">✓ Currently Allowed Students ({{ allowed_list|length }} total)</h4>
                    <div style="max-height: 300px; overflow-y: auto;">
                        <table>
                            <thead>
                                <tr style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white;">
                                    <th style="color: white;">#</th>
                                    <th style="color: white;">Student Name</th>
                                    <th style="color: white;">Student ID</th>
                                </tr>
                            </thead>
                            <tbody>
                                {% for s in allowed_list %}
                                    <tr style="border-bottom: 1px solid #e0e0e0;">
                                        <td style="color: #667eea; font-weight: 600;">{{ loop.index }}</td>
                                        <td><strong>{{ s.name }}</strong></td>
                                        <td><code style="background: #f0f0f0; padding: 3px 8px; border-radius: 4px;">{{ s.student_id }}</code></td>
                                    </tr>
                                {% endfor %}
                            </tbody>
                        </table>
                    </div>
                </div>
            {% else %}
                <div style="margin-top: 30px; background: #fff3e0; padding: 20px; border-radius: 8px; border-left: 4px solid #ffc107;">
                    <strong style="color: #e8840c;">📋 No students uploaded yet</strong>
                    <p style="color: #666; margin: 8px 0 0 0;">Upload a CSV file above to specify which students can take this quiz.</p>
                </div>
            {% endif %}

            <hr style="margin: 30px 0;">
            <div class="alert-box error">
                ⚠️ Danger Zone
            </div>
            <form method="POST" action="{{ url_for('clear_results') }}" onsubmit="return confirm('🚨 Delete ALL student results? This cannot be undone!');">
                <button type="submit" class="btn btn-primary" style="background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);">🗑️ Delete All Results</button>
            </form>
        </div>

        <div id="results-section" class="content-panel">
            <h2>📊 Student Results</h2>
            <hr style="margin: 20px 0;">
            
            {% if student_results %}
                <p style="color: #666; margin-bottom: 20px;">Total Submissions: <strong>{{ student_results|length }}</strong></p>
                <table>
                    <thead>
                        <tr>
                            <th>Student Name</th>
                            <th style="text-align: center;">Score</th>
                            <th style="text-align: center;">Percentage</th>
                            <th style="text-align: center;">Date & Time</th>
                            <th style="text-align: center;">Action</th>
                        </tr>
                    </thead>
                    <tbody>
                        {% for result in student_results %}
                            <tr>
                                <td><strong>{{ result.student_name }}</strong></td>
                                <td style="text-align: center;">{{ result.score }}/{{ result.total }}</td>
                                <td style="text-align: center;">
                                    <span style="display: inline-block; padding: 4px 12px; border-radius: 20px; font-weight: 600; color: white; background: {% if result.percentage >= 70 %}#27ae60{% elif result.percentage >= 50 %}#f39c12{% else %}#e74c3c{% endif %};">
                                        {{ result.percentage }}%
                                    </span>
                                </td>
                                <td style="text-align: center; font-size: 0.9rem;">{{ result.date }}<br>{{ result.time }}</td>
                                <td style="text-align: center;">
                                    <a href="{{ url_for('download_pdf', index=loop.index0) }}" class="btn btn-primary" style="padding: 8px 16px; font-size: 0.9rem;">📥 PDF</a>
                                </td>
                            </tr>
                        {% endfor %}
                    </tbody>
                </table>

                <div style="margin-top: 30px; padding: 20px; background: linear-gradient(135deg, #e3f2fd 0%, #bbdefb 100%); border-radius: 8px; text-align: center;">
                    <h4>📥 Download All Results</h4>
                    <p style="color: #1565c0; margin: 10px 0;">Export to Excel spreadsheet</p>
                    <a href="{{ url_for('download_excel') }}" class="btn btn-info">📊 Download Excel</a>
                </div>
            {% else %}
                <div class="alert-box info">
                    📋 No student results yet. Results will appear here once students submit quizzes.
                </div>
            {% endif %}
        </div>

        <div id="analytics-section" class="content-panel">
            <h2>📈 Performance Analytics</h2>
            <hr style="margin: 20px 0;">
            
            {% if analytics.total_students > 0 %}
                <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(250px, 1fr)); gap: 20px; margin-bottom: 30px;">
                    <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 25px; border-radius: 12px;">
                        <div style="font-size: 0.9rem; opacity: 0.9;">Total Students</div>
                        <div style="font-size: 2.5rem; font-weight: 700; margin-top: 10px;">{{ analytics.total_students }}</div>
                    </div>
                    <div style="background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%); color: white; padding: 25px; border-radius: 12px;">
                        <div style="font-size: 0.9rem; opacity: 0.9;">Average Score</div>
                        <div style="font-size: 2.5rem; font-weight: 700; margin-top: 10px;">{{ analytics.average_score }}%</div>
                    </div>
                    <div style="background: linear-gradient(135deg, #4facfe 0%, #00f2fe 100%); color: white; padding: 25px; border-radius: 12px;">
                        <div style="font-size: 0.9rem; opacity: 0.9;">Pass Rate</div>
                        <div style="font-size: 2.5rem; font-weight: 700; margin-top: 10px;">{{ analytics.pass_rate }}%</div>
                    </div>
                    <div style="background: linear-gradient(135deg, #43e97b 0%, #38f9d7 100%); color: white; padding: 25px; border-radius: 12px;">
                        <div style="font-size: 0.9rem; opacity: 0.9;">Highest Score</div>
                        <div style="font-size: 2.5rem; font-weight: 700; margin-top: 10px;">{{ analytics.highest_score }}%</div>
                    </div>
                </div>

                <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(350px, 1fr)); gap: 20px; margin-bottom: 30px;">
                    <div class="chart-container">
                        <h4>Score Distribution</h4>
                        <canvas id="scoreChart"></canvas>
                    </div>
                    <div class="chart-container">
                        <h4>Pass vs Fail</h4>
                        <canvas id="passFailChart"></canvas>
                    </div>
                </div>

                <div class="chart-container">
                    <h4>Detailed Statistics</h4>
                    <table>
                        <tr>
                            <td><strong>Total Questions:</strong></td>
                            <td>{{ questions|length }}</td>
                            <td><strong>Highest Score:</strong></td>
                            <td><span style="color: #27ae60; font-weight: 700;">{{ analytics.highest_score }}%</span></td>
                        </tr>
                        <tr>
                            <td><strong>Passed (≥70%):</strong></td>
                            <td><span style="color: #27ae60; font-weight: 700;">{{ analytics.pass_count }}</span></td>
                            <td><strong>Lowest Score:</strong></td>
                            <td><span style="color: #e74c3c; font-weight: 700;">{{ analytics.lowest_score }}%</span></td>
                        </tr>
                        <tr>
                            <td><strong>Failed (<70%):</strong></td>
                            <td><span style="color: #e74c3c; font-weight: 700;">{{ analytics.fail_count }}</span></td>
                            <td><strong>Std. Deviation:</strong></td>
                            <td>{{ analytics.std_dev|default('N/A') }}</td>
                        </tr>
                    </table>
                </div>

                <script>
                    const percentages = {{ percentages|tojson }};
                    if (percentages && percentages.length > 0) {
                        const gradeRanges = {
                            'A (90-100%)': percentages.filter(p => p >= 90).length,
                            'B (80-89%)': percentages.filter(p => p >= 80 && p < 90).length,
                            'C (70-79%)': percentages.filter(p => p >= 70 && p < 80).length,
                            'D (60-69%)': percentages.filter(p => p >= 60 && p < 70).length,
                            'F (<60%)': percentages.filter(p => p < 60).length
                        };
                        
                        const ctx1 = document.getElementById('scoreChart').getContext('2d');
                        new Chart(ctx1, {
                            type: 'bar',
                            data: {
                                labels: Object.keys(gradeRanges),
                                datasets: [{
                                    label: 'Students',
                                    data: Object.values(gradeRanges),
                                    backgroundColor: ['#27ae60', '#f39c12', '#3498db', '#e67e22', '#e74c3c'],
                                    borderRadius: 8,
                                    borderSkipped: false
                                }]
                            },
                            options: { responsive: true, plugins: { legend: { display: false } }, scales: { y: { beginAtZero: true } } }
                        });
                        
                        const ctx2 = document.getElementById('passFailChart').getContext('2d');
                        new Chart(ctx2, {
                            type: 'doughnut',
                            data: {
                                labels: ['Passed ≥70%', 'Failed <70%'],
                                datasets: [{
                                    data: [{{ analytics.pass_count }}, {{ analytics.fail_count }}],
                                    backgroundColor: ['#27ae60', '#e74c3c'],
                                    borderColor: '#fff',
                                    borderWidth: 3
                                }]
                            },
                            options: { responsive: true, plugins: { legend: { position: 'bottom' } } }
                        });
                    }
                </script>
            {% else %}
                <div class="alert-box info">
                    📊 Analytics will appear here once students complete quizzes.
                </div>
            {% endif %}
        </div>
    </div>
</body>
</html>
'''

QUIZ_SELECT_TEMPLATE = '''
<!DOCTYPE html>
<html>
<head>
    <title>Select Quiz</title>
    <link rel="icon" type="image/svg+xml" href="{{ url_for('static', filename='favicon.svg') }}">
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 40px 20px;
        }
        .quiz-container {
            max-width: 900px;
            margin: 0 auto;
        }
        .page-header {
            text-align: center;
            color: white;
            margin-bottom: 50px;
            animation: slideDown 0.6s ease;
        }
        @keyframes slideDown {
            from {
                opacity: 0;
                transform: translateY(-20px);
            }
            to {
                opacity: 1;
                transform: translateY(0);
            }
        }
        .page-header h1 {
            font-size: 2.5rem;
            font-weight: 700;
            margin-bottom: 10px;
        }
        .page-header p {
            font-size: 1.1rem;
            opacity: 0.9;
        }
        .quiz-card {
            background: white;
            padding: 25px;
            margin: 20px 0;
            border-radius: 15px;
            box-shadow: 0 10px 30px rgba(0, 0, 0, 0.2);
            display: flex;
            justify-content: space-between;
            align-items: center;
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
            animation: slideUp 0.5s ease forwards;
            opacity: 0;
        }
        @keyframes slideUp {
            from {
                opacity: 0;
                transform: translateY(20px);
            }
            to {
                opacity: 1;
                transform: translateY(0);
            }
        }
        {% for quiz in available_quizzes %}
            .quiz-card:nth-child({{ loop.index }}) {
                animation-delay: {{ loop.index0 * 0.1 }}s;
            }
        {% endfor %}
        .quiz-card:hover {
            transform: translateY(-5px);
            box-shadow: 0 15px 40px rgba(0, 0, 0, 0.3);
        }
        .quiz-info {
            flex: 1;
            margin-right: 20px;
        }
        .quiz-info h3 {
            color: #333;
            font-weight: 700;
            font-size: 1.3rem;
            margin-bottom: 10px;
        }
        .quiz-meta {
            display: flex;
            gap: 20px;
            margin-bottom: 12px;
            flex-wrap: wrap;
        }
        .meta-item {
            color: #666;
            display: flex;
            align-items: center;
            gap: 6px;
        }
        .timer-badge {
            display: inline-block;
            background: linear-gradient(135deg, #ffc107 0%, #ff9800 100%);
            color: white;
            padding: 8px 16px;
            border-radius: 20px;
            font-size: 0.9rem;
            font-weight: 600;
        }
        .start-btn {
            padding: 12px 30px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            text-decoration: none;
            border-radius: 10px;
            font-weight: 600;
            transition: all 0.3s ease;
            white-space: nowrap;
            display: inline-block;
        }
        .start-btn:hover {
            transform: translateY(-2px);
            box-shadow: 0 8px 20px rgba(102, 126, 234, 0.4);
            color: white;
        }
        .no-quizzes {
            background: white;
            padding: 60px 40px;
            text-align: center;
            border-radius: 15px;
            box-shadow: 0 10px 30px rgba(0, 0, 0, 0.2);
        }
        .no-quizzes h3 {
            color: #333;
            font-weight: 700;
            margin-bottom: 10px;
        }
        .no-quizzes p {
            color: #666;
            font-size: 1.1rem;
        }
        .back-btn {
            display: inline-block;
            margin-top: 30px;
            padding: 12px 30px;
            background: rgba(255, 255, 255, 0.2);
            color: white;
            text-decoration: none;
            border-radius: 10px;
            font-weight: 600;
            transition: all 0.3s ease;
            border: 2px solid white;
        }
        .back-btn:hover {
            background: rgba(255, 255, 255, 0.3);
            color: white;
        }
    </style>
</head>
<body>
    <div class="quiz-container">
        <div class="page-header">
            <h1>📚 Available Quizzes</h1>
            <p>Choose a quiz to get started</p>
        </div>
        
        {% if available_quizzes %}
            {% for quiz in available_quizzes %}
                <div class="quiz-card">
                    <div class="quiz-info">
                        <h3>{{ quiz.admin_name }}'s Quiz</h3>
                        <div class="quiz-meta">
                            <div class="meta-item">📝 {{ quiz.question_count }} Question{{ 's' if quiz.question_count != 1 else '' }}</div>
                            {% if quiz.time_limit > 0 %}
                                <div class="meta-item">⏱️ {{ quiz.time_limit }} min{{ 's' if quiz.time_limit != 1 else '' }}</div>
                            {% else %}
                                <div class="meta-item">♾️ Unlimited Time</div>
                            {% endif %}
                        </div>
                        {% if quiz.time_limit > 0 %}
                            <span class="timer-badge">⏱️ {{ quiz.time_limit }} minutes</span>
                        {% else %}
                            <span class="timer-badge" style="background: linear-gradient(135deg, #4facfe 0%, #00f2fe 100%);">No Time Limit</span>
                        {% endif %}
                    </div>
                    <a href="{{ url_for('take_quiz', admin_id=quiz.admin_id) }}" class="start-btn">Start Quiz →</a>
                </div>
            {% endfor %}
        {% else %}
            <div class="no-quizzes">
                <h3>No Quizzes Available</h3>
                <p>Instructors haven't created any quizzes yet. Please check back later!</p>
                <a href="{{ url_for('home') }}" class="back-btn">← Back to Home</a>
            </div>
        {% endif %}
        
        {% if available_quizzes %}
            <div style="text-align: center; margin-top: 40px;">
                <a href="{{ url_for('home') }}" class="back-btn">← Back to Home</a>
            </div>
        {% endif %}
    </div>
</body>
</html>
'''


USER_QUIZ_TEMPLATE = '''
<!DOCTYPE html>
<html>
<head>
    <title>Take Quiz</title>
    <link rel="icon" type="image/svg+xml" href="{{ url_for('static', filename='favicon.svg') }}">
    <style>
        body {
            font-family: Arial, sans-serif;
            max-width: 800px;
            margin: 20px auto;
            padding: 20px;
            background: linear-gradient(135deg, #3498db 0%, #2980b9 100%);
            min-height: 100vh;
        }
        .container {
            background: white;
            padding: 40px;
            border-radius: 10px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        }
        h2 {
            color: #333;
            text-align: center;
            margin-bottom: 30px;
        }
        .quiz-info {
            background: #f0f0f0;
            padding: 15px;
            border-radius: 5px;
            margin-bottom: 20px;
            text-align: center;
        }
        .timer-display {
            background: #fff3cd;
            border: 3px solid #ffc107;
            padding: 20px;
            border-radius: 10px;
            margin: 20px 0;
            text-align: center;
            position: sticky;
            top: 20px;
            z-index: 1000;
        }
        .timer-display.warning {
            background: #f8d7da;
            border-color: #e74c3c;
            animation: pulse 1s infinite;
        }
        @keyframes pulse {
            0%, 100% { opacity: 1; }
            50% { opacity: 0.7; }
        }
        .timer-text {
            font-size: 36px;
            font-weight: bold;
            color: #333;
            margin: 10px 0;
        }
        .timer-text.warning {
            color: #e74c3c;
        }
        .question {
            background: #f9f9f9;
            padding: 20px;
            margin: 20px 0;
            border-radius: 5px;
            border-left: 4px solid #3498db;
        }
        .question h3 {
            color: #333;
            margin-bottom: 15px;
        }
        .options {
            margin: 10px 0;
        }
        .option {
            padding: 10px;
            margin: 8px 0;
            cursor: pointer;
            border: 2px solid #ddd;
            border-radius: 5px;
            transition: all 0.3s;
        }
        .option:hover {
            background: #e8f4f8;
            border-color: #3498db;
        }
        .option input {
            margin-right: 10px;
        }
        .btn {
            width: 100%;
            padding: 15px;
            background: #3498db;
            color: white;
            border: none;
            border-radius: 5px;
            cursor: pointer;
            font-size: 18px;
            margin-top: 20px;
        }
        .btn:hover {
            background: #2980b9;
        }
        .back-link {
            display: block;
            text-align: center;
            margin-top: 20px;
            color: #3498db;
            text-decoration: none;
        }
        .name-input {
            width: 100%;
            padding: 12px;
            margin: 20px 0;
            border: 2px solid #ddd;
            border-radius: 5px;
            font-size: 16px;
        }
        .no-questions {
            text-align: center;
            padding: 40px;
            color: #888;
        }
    </style>
    <script>
        let timeLimit = {{ time_limit }};  // in minutes
        let timeRemaining = timeLimit * 60;  // convert to seconds
        
        function updateTimer() {
            if (timeLimit === 0) return;  // No timer if time limit is 0
            
            const minutes = Math.floor(timeRemaining / 60);
            const seconds = timeRemaining % 60;
            
            const timerDisplay = document.getElementById('timer-display');
            const timerText = document.getElementById('timer-text');
            
            timerText.textContent = `${minutes}:${seconds.toString().padStart(2, '0')}`;
            
            // Warning when less than 5 minutes
            if (timeRemaining <= 300) {
                timerDisplay.classList.add('warning');
                timerText.classList.add('warning');
            }
            
            if (timeRemaining <= 0) {
                // Auto-submit when time runs out
                alert('Time is up! Your quiz will be submitted automatically.');
                document.getElementById('quiz-form').submit();
            }
            
            timeRemaining--;
        }
        
        window.onload = function() {
            if (timeLimit > 0) {
                updateTimer();  // Initial call
                setInterval(updateTimer, 1000);  // Update every second
            }
        };
    </script>
</head>
<body>
    <div class="container">
        <h2>✏️ Quiz Time!</h2>
        
        {% with messages = get_flashed_messages(with_categories=true) %}
            {% if messages %}
                <div style="margin-bottom:15px;">
                    {% for category, msg in messages %}
                        <div style="padding:12px; border-radius:6px; margin-bottom:6px; background: {% if category == 'error' %}#fdecea{% elif category == 'success' %}#e8f5e9{% else %}#eef2ff{% endif %}; color: {% if category == 'error' %}#611a15{% elif category == 'success' %}#1b5e20{% else %}#1e3a8a{% endif %};">
                            {{ msg }}
                        </div>
                    {% endfor %}
                </div>
            {% endif %}
        {% endwith %}

        {% if questions %}
            <div class="quiz-info">
                <strong>Quiz by:</strong> {{ admin_name }} | <strong>Questions:</strong> {{ questions|length }}
            </div>
            
            {% if time_limit > 0 %}
                <div id="timer-display" class="timer-display" style="display:none;">
                    <div style="font-size: 18px; color: #666;">⏱️ Time Remaining</div>
                    <div id="timer-text" class="timer-text">{{ time_limit }}:00</div>
                    <div style="font-size: 14px; color: #888;">Your quiz will auto-submit when time runs out</div>
                </div>
            {% endif %}

            <!-- Initial auth block -->
            <div id="auth-block" style="background:#f7f9fb;padding:20px;border-radius:8px;margin-bottom:16px;">
                <h3>Enter your details to start</h3>
                <input type="text" id="auth-name" class="name-input" placeholder="Full name" required>
                <input type="text" id="auth-id" class="name-input" placeholder="Registration / Reg No" required>
                <div id="auth-warning" style="color:#e74c3c; display:none; margin-top:8px;"></div>
                <button id="start-btn" class="btn" style="background:#27ae60; margin-top:10px;">Start Quiz</button>
            </div>

            <form id="quiz-form" method="POST" action="{{ url_for('submit_quiz') }}" style="display:none;">
                <input type="hidden" id="student-name-hidden" name="student_name">
                <input type="hidden" id="student-id-hidden" name="student_id">
                <input type="hidden" name="admin_id" value="{{ admin_id }}">

                <div id="quiz-content">
                {% for q in questions %}
                    <div class="question">
                        <h3>Question {{ loop.index }}: {{ q.question }}</h3>
                        <div class="options">
                            <label class="option">
                                <input type="radio" name="q{{ loop.index0 }}" value="A" required>
                                A) {{ q.options.A }}
                            </label>
                            <label class="option">
                                <input type="radio" name="q{{ loop.index0 }}" value="B">
                                B) {{ q.options.B }}
                            </label>
                            <label class="option">
                                <input type="radio" name="q{{ loop.index0 }}" value="C">
                                C) {{ q.options.C }}
                            </label>
                            <label class="option">
                                <input type="radio" name="q{{ loop.index0 }}" value="D">
                                D) {{ q.options.D }}
                            </label>
                        </div>
                    </div>
                {% endfor %}
                </div>

                <button type="submit" id="submit-btn" class="btn">Submit Quiz</button>
            </form>

            <script>
                // Client-side entry gate: require name + reg no to be in allowed list
                const allowedEntries = {{ allowed_list|tojson }} || [];
                const existingNames = {{ existing_names|tojson }} || [];

                const authName = document.getElementById('auth-name');
                const authId = document.getElementById('auth-id');
                const authWarning = document.getElementById('auth-warning');
                const startBtn = document.getElementById('start-btn');
                const quizForm = document.getElementById('quiz-form');
                const studentNameHidden = document.getElementById('student-name-hidden');
                const studentIdHidden = document.getElementById('student-id-hidden');
                const timerDisplay = document.getElementById('timer-display');

                function startQuiz() {
                    const nameVal = (authName.value || '').trim();
                    const idVal = (authId.value || '').trim();
                    if (!nameVal || !idVal) {
                        authWarning.textContent = 'Please enter both name and registration number.';
                        authWarning.style.display = 'block';
                        return;
                    }

                    const allowed = allowedEntries.some(e => (e.name||'').trim().toLowerCase() === nameVal.toLowerCase() && (e.student_id||'').trim().toLowerCase() === idVal.toLowerCase());
                    if (!allowed) {
                        authWarning.textContent = 'You are not allowed to take this quiz. Please contact the instructor.';
                        authWarning.style.display = 'block';
                        return;
                    }

                    const already = existingNames.some(n => (n||'').trim().toLowerCase() === nameVal.toLowerCase());
                    if (already) {
                        authWarning.textContent = 'You have already submitted this quiz.';
                        authWarning.style.display = 'block';
                        return;
                    }

                    // Passed checks: populate hidden fields and show quiz
                    studentNameHidden.value = nameVal;
                    studentIdHidden.value = idVal;
                    quizForm.style.display = 'block';
                    document.getElementById('auth-block').style.display = 'none';
                    if (timerDisplay) timerDisplay.style.display = 'block';
                    // Scroll to quiz
                    quizForm.scrollIntoView({behavior: 'smooth'});
                }

                startBtn.addEventListener('click', function(e){ e.preventDefault(); startQuiz(); });
            </script>
        {% else %}
            <div class="no-questions">
                <h3>No questions available yet!</h3>
                <p>This admin hasn't added questions yet.</p>
            </div>
        {% endif %}
        
        <a href="{{ url_for('user_quiz') }}" class="back-link">← Back to Quiz Selection</a>
    </div>
</body>
</html>
'''

RESULTS_TEMPLATE = '''
<!DOCTYPE html>
<html>
<head>
    <title>Quiz Results</title>
    <link rel="icon" type="image/svg+xml" href="{{ url_for('static', filename='favicon.svg') }}">
    <style>
        body {
            font-family: Arial, sans-serif;
            max-width: 800px;
            margin: 20px auto;
            padding: 20px;
            background: linear-gradient(135deg, #27ae60 0%, #229954 100%);
            min-height: 100vh;
        }
        .container {
            background: white;
            padding: 40px;
            border-radius: 10px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        }
        h2 {
            color: #333;
            text-align: center;
        }
        .score {
            text-align: center;
            font-size: 48px;
            color: #27ae60;
            margin: 30px 0;
        }
        .result-item {
            background: #f9f9f9;
            padding: 15px;
            margin: 15px 0;
            border-radius: 5px;
        }
        .correct {
            border-left: 4px solid #27ae60;
        }
        .incorrect {
            border-left: 4px solid #e74c3c;
        }
        .question-text {
            font-weight: bold;
            margin-bottom: 10px;
        }
        .answer-comparison {
            margin: 5px 0;
        }
        .your-answer {
            color: #e74c3c;
        }
        .correct-answer {
            color: #27ae60;
        }
        .btn {
            display: block;
            width: 200px;
            margin: 20px auto;
            padding: 12px;
            background: #3498db;
            color: white;
            text-align: center;
            text-decoration: none;
            border-radius: 5px;
        }
        .btn:hover {
            background: #2980b9;
        }
    </style>
</head>
<body>
    <div class="container">
        <h2>🎉 Quiz Results</h2>
        <h3 style="text-align: center;">Student: {{ student_name }}</h3>
        <div class="score">{{ score }} / {{ total }}</div>
        <p style="text-align: center; font-size: 20px;">
            Percentage: {{ percentage }}%
        </p>
        
        <h3>Detailed Results:</h3>
        {% for item in results %}
            <div class="result-item {% if item.correct %}correct{% else %}incorrect{% endif %}">
                <div class="question-text">Q{{ loop.index }}: {{ item.question }}</div>
                <div class="answer-comparison">
                    <span class="your-answer">Your answer: {{ item.user_answer }}</span>
                </div>
                {% if not item.correct %}
                    <div class="answer-comparison">
                        <span class="correct-answer">Correct answer: {{ item.correct_answer }}</span>
                    </div>
                {% endif %}
            </div>
        {% endfor %}
        
        <a href="{{ url_for('home') }}" class="btn">Back to Home</a>
    </div>
</body>
</html>
'''

# Routes
@app.route('/')
def home():
    init_files()
    return render_template_string(HOME_TEMPLATE)

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    # Clean up inactive admins on each login attempt
    cleanup_inactive_admins()
    
    if request.method == 'POST':
        action = request.form.get('action')
        username = request.form.get('username')
        password = request.form.get('password')
        
        admins = load_admins()
        
        if action == 'register':
            email = request.form.get('email', '').strip()
            phone = request.form.get('phone', '').strip()
            confirm_password = request.form.get('confirm_password')
            
            if not email:
                flash('Please provide an email address!', 'error')
                return redirect(url_for('admin_login'))
            
            if not phone:
                flash('Please provide a phone number!', 'error')
                return redirect(url_for('admin_login'))
            
            if password != confirm_password:
                flash('Passwords do not match!', 'error')
                return redirect(url_for('admin_login'))
            
            if username in admins:
                flash('Username already exists!', 'error')
                return redirect(url_for('admin_login'))
            
            # Create new admin with email and phone
            admins[username] = {
                'password': password,
                'email': email,
                'phone': phone,
                'created_at': datetime.now().isoformat()
            }
            save_admins(admins)
            
            # Initialize quiz settings for new admin
            quiz_settings = load_quiz_settings()
            quiz_settings[username] = {'time_limit': 0}  # Default: no time limit
            save_quiz_settings(quiz_settings)
            
            # Send confirmation email
            try:
                send_admin_welcome_email(username, email)
            except Exception as e:
                print(f'Failed to send email: {e}')
            
            flash('Admin account created successfully! A confirmation email has been sent. Please login.', 'success')
            return redirect(url_for('admin_login'))
        
        elif action == 'login':
            if username in admins and admins[username]['password'] == password:
                # Make session permanent so it uses `permanent_session_lifetime`
                session.permanent = True
                session['admin'] = username
                session['last_active'] = datetime.utcnow().isoformat()
                return redirect(url_for('admin_panel'))
            else:
                flash('Invalid username or password!', 'error')
    
    return render_template_string(ADMIN_LOGIN_TEMPLATE)


@app.route('/start', methods=['GET', 'POST'])
def start():
    if request.method == 'POST':
        student_name = request.form.get('student_name', '').strip()
        student_id = request.form.get('student_id', '').strip()
        if not student_name or not student_id:
            flash('Please enter both name and registration number.', 'error')
            return render_template_string(START_TEMPLATE)

        # Find allowed admins where this student appears
        allowed = load_allowed()
        matches = []
        for admin_id, lst in allowed.items():
            for entry in lst:
                if (entry.get('name') or '').strip().lower() == student_name.lower() and (entry.get('student_id') or '').strip().lower() == student_id.lower():
                    matches.append(admin_id)
                    break

        if not matches:
            flash('You are not allowed to take any quizzes. Contact the instructor.', 'error')
            return render_template_string(START_TEMPLATE)

        # Save in session and redirect to quiz selection (filtered)
        session['student_name'] = student_name
        session['student_id'] = student_id
        session['allowed_admins'] = matches
        return redirect(url_for('user_quiz'))

    return render_template_string(START_TEMPLATE)


@app.route('/student_logout')
def student_logout():
    session.pop('student_name', None)
    session.pop('student_id', None)
    session.pop('allowed_admins', None)
    return redirect(url_for('start'))

@app.route('/admin/panel', methods=['GET', 'POST'])
@login_required
def admin_panel():
    if 'admin' not in session:
        return redirect(url_for('admin_login'))
    
    current_admin = session['admin']
    all_questions = load_questions()
    
    # Get questions for current admin only
    admin_questions = all_questions.get(current_admin, [])
    
    if request.method == 'POST':
        question = {
            'question': request.form.get('question'),
            'options': {
                'A': request.form.get('option_a'),
                'B': request.form.get('option_b'),
                'C': request.form.get('option_c'),
                'D': request.form.get('option_d')
            },
            'correct_answer': request.form.get('correct_answer')
        }
        
        admin_questions.append(question)
        all_questions[current_admin] = admin_questions
        save_questions(all_questions)
        
        return redirect(url_for('admin_panel'))
    
    # Get answers for current admin's quizzes
    all_answers = load_answers()
    admin_answers = all_answers.get(current_admin, [])

    # Unique submitted student names (preserve order)
    submitted_names = []
    seen = set()
    for entry in admin_answers:
        name = (entry.get('student_name') or '').strip()
        key = name.lower()
        if name and key not in seen:
            submitted_names.append(name)
            seen.add(key)
    
    # Get quiz settings
    quiz_settings = load_quiz_settings()
    admin_settings = quiz_settings.get(current_admin, {'time_limit': 0})
    
    # Process student results for display
    student_results = []
    percentages = []
    for answer in admin_answers:
        timestamp = datetime.fromisoformat(answer['timestamp'])
        percentage = round((answer['score'] / answer['total']) * 100, 2) if answer['total'] > 0 else 0
        percentages.append(percentage)
        student_results.append({
            'student_name': answer['student_name'],
            'score': answer['score'],
            'total': answer['total'],
            'percentage': percentage,
            'date': timestamp.strftime('%Y-%m-%d'),
            'time': timestamp.strftime('%I:%M %p')
        })
    
    # Calculate analytics
    analytics = {
        'total_students': len(student_results),
        'average_score': round(sum(percentages) / len(percentages), 2) if percentages else 0,
        'highest_score': max(percentages) if percentages else 0,
        'lowest_score': min(percentages) if percentages else 0,
        'pass_count': sum(1 for p in percentages if p >= 70) if percentages else 0,
        'fail_count': sum(1 for p in percentages if p < 70) if percentages else 0,
        'pass_rate': round((sum(1 for p in percentages if p >= 70) / len(percentages) * 100), 2) if percentages else 0
    }
    # Load allowed students for this admin
    allowed_data = load_allowed()
    allowed_list = allowed_data.get(current_admin, [])

    return render_template_string(ADMIN_PANEL_TEMPLATE, 
                                 current_admin=current_admin,
                                 questions=admin_questions,
                                 student_results=student_results,
                                 quiz_time_limit=admin_settings['time_limit'],
                                 analytics=analytics,
                                 percentages=percentages,
                                 submitted_names=submitted_names,
                                 allowed_list=allowed_list)

@app.route('/admin/update-settings', methods=['POST'])
@login_required
def update_quiz_settings():
    if 'admin' not in session:
        return redirect(url_for('admin_login'))
    
    current_admin = session['admin']
    time_limit = int(request.form.get('time_limit', 0))
    
    quiz_settings = load_quiz_settings()
    quiz_settings[current_admin] = {'time_limit': time_limit}
    save_quiz_settings(quiz_settings)
    
    flash('Quiz timer settings updated successfully!', 'success')
    return redirect(url_for('admin_panel'))


@app.route('/admin/upload-students', methods=['POST'])
@login_required
def upload_students():
    if 'admin' not in session:
        return redirect(url_for('admin_login'))
    current_admin = session['admin']

    if 'students_file' not in request.files:
        flash('No file uploaded.', 'error')
        return redirect(url_for('admin_panel'))

    f = request.files['students_file']
    if f.filename == '':
        flash('No file selected.', 'error')
        return redirect(url_for('admin_panel'))

    try:
        raw_content = f.stream.read().decode('utf-8').strip()
        lines = raw_content.splitlines()
        
        if not lines:
            flash('CSV file is empty.', 'error')
            return redirect(url_for('admin_panel'))
        
        # Detect delimiter: check first line for tabs or commas
        first_line = lines[0]
        delimiter = '\t' if '\t' in first_line else ','
        
        # Parse CSV with detected delimiter
        reader = csv.reader(lines, delimiter=delimiter)
        headers = next(reader, None)
        
        if not headers:
            flash('CSV file has no headers.', 'error')
            return redirect(url_for('admin_panel'))
        
        # Find column indices for name and student_id
        headers_lower = [h.lower().strip() for h in headers]
        name_col = None
        sid_col = None
        
        for i, h in enumerate(headers_lower):
            if 'name' in h or 'student' in h:
                if 'id' not in h:
                    name_col = i
            if 'id' in h or 'studentid' in h or 'reg' in h:
                sid_col = i
        
        # Fallback: assume first column is name, second is ID
        if name_col is None:
            name_col = 0
        if sid_col is None:
            sid_col = 1 if len(headers) > 1 else None
        
        new_list = []
        for row_num, row in enumerate(reader, start=2):
            try:
                name = (row[name_col].strip() if name_col < len(row) else '').strip()
                sid = (row[sid_col].strip() if sid_col and sid_col < len(row) else '').strip()
                
                if name and sid:
                    new_list.append({'name': name, 'student_id': sid})
            except (IndexError, ValueError):
                pass
        
        if not new_list:
            flash('No valid student entries found in CSV. Ensure format: Name, Student ID', 'error')
            return redirect(url_for('admin_panel'))

        # Save to allowed file under current admin
        allowed = load_allowed()
        allowed[current_admin] = new_list
        save_allowed(allowed)
        
        flash(f'✓ Successfully uploaded {len(new_list)} student(s)!', 'success')
    except Exception as e:
        flash(f'Failed to process CSV: {str(e)}. Ensure format is: Name (first column), Student ID (second column)', 'error')

    return redirect(url_for('admin_panel'))


@app.route('/admin/clear-results', methods=['POST'])
@login_required
def clear_results():
    if 'admin' not in session:
        return redirect(url_for('admin_login'))
    current_admin = session['admin']
    # Clear only current admin's results
    all_answers = load_answers()
    if current_admin in all_answers:
        all_answers[current_admin] = []
        save_answers(all_answers)
        flash('All student results cleared for this admin.', 'success')
    else:
        flash('No student results to clear.', 'error')
    return redirect(url_for('admin_panel'))

@app.route('/admin/delete/<int:index>', methods=['POST'])
@login_required
def delete_question(index):
    if 'admin' not in session:
        return redirect(url_for('admin_login'))
    
    current_admin = session['admin']
    all_questions = load_questions()
    admin_questions = all_questions.get(current_admin, [])
    
    if 0 <= index < len(admin_questions):
        admin_questions.pop(index)
        all_questions[current_admin] = admin_questions
        save_questions(all_questions)
    
    return redirect(url_for('admin_panel'))

@app.route('/admin/download-pdf/<int:index>')
@login_required
def download_pdf(index):
    if 'admin' not in session:
        return redirect(url_for('admin_login'))
    
    current_admin = session['admin']
    all_answers = load_answers()
    admin_answers = all_answers.get(current_admin, [])
    
    if 0 <= index < len(admin_answers):
        result = admin_answers[index]
        timestamp = datetime.fromisoformat(result['timestamp'])
        percentage = round((result['score'] / result['total']) * 100, 2) if result['total'] > 0 else 0
        
        # Create PDF in memory
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        
        # Container for PDF elements
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#333333'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Title
        title = Paragraph("Quiz Result Report", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.2*inch))
        
        # Student Info
        student_info_data = [
            ['Student Name:', result['student_name']],
            ['Date:', timestamp.strftime('%Y-%m-%d')],
            ['Time:', timestamp.strftime('%I:%M %p')],
        ]
        
        student_table = Table(student_info_data, colWidths=[2*inch, 4*inch])
        student_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f0f0f0')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('TOPPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#dddddd'))
        ]))
        elements.append(student_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Score Summary
        grade = 'A' if percentage >= 70 else 'B' if percentage >= 60 else 'C' if percentage >= 50 else 'D' if percentage >= 40 else 'F'
        
        score_data = [
            ['SCORE', 'PERCENTAGE', 'GRADE'],
            [f"{result['score']}/{result['total']}", f"{percentage}%", grade]
        ]
        
        score_table = Table(score_data, colWidths=[2*inch, 2*inch, 2*inch])
        score_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e74c3c')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('FONTSIZE', (0, 1), (-1, 1), 18),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 15),
            ('TOPPADDING', (0, 0), (-1, -1), 15),
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#f9f9f9')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#dddddd'))
        ]))
        elements.append(score_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Questions
        for idx, item in enumerate(result['results'], 1):
            question_data = [
                [f"Q{idx}: {item['question']}"],
                [f"Student's Answer: {item['user_answer']}"]
            ]
            
            if not item['correct']:
                question_data.append([f"Correct Answer: {item['correct_answer']}"])
            
            question_table = Table(question_data, colWidths=[6.5*inch])
            
            table_style = [
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f9f9f9')),
                ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('PADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#dddddd')),
            ]
            
            if item['correct']:
                table_style.append(('LINEAFTER', (0, 0), (0, -1), 4, colors.HexColor('#27ae60')))
            else:
                table_style.append(('LINEAFTER', (0, 0), (0, -1), 4, colors.HexColor('#e74c3c')))
            
            question_table.setStyle(TableStyle(table_style))
            elements.append(question_table)
            elements.append(Spacer(1, 0.15*inch))
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        # Return PDF as download
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"Quiz_Result_{result['student_name'].replace(' ', '_')}_{timestamp.strftime('%Y%m%d')}.pdf",
            mimetype='application/pdf'
        )
    
    return redirect(url_for('admin_panel'))

@app.route('/admin/download-excel')
@login_required
def download_excel():
    if 'admin' not in session:
        return redirect(url_for('admin_login'))
    
    current_admin = session['admin']
    all_answers = load_answers()
    admin_answers = all_answers.get(current_admin, [])
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Results"
    
    # Style definitions
    header_fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Headers
    headers = ['Student Name', 'Score', 'Total', 'Percentage', 'Grade', 'Date', 'Time']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for row_idx, answer in enumerate(admin_answers, 2):
        timestamp = datetime.fromisoformat(answer['timestamp'])
        percentage = round((answer['score'] / answer['total']) * 100, 2) if answer['total'] > 0 else 0
        grade = 'A' if percentage >= 70 else 'B' if percentage >= 60 else 'C' if percentage >= 50 else 'D' if percentage >= 40 else 'F'
        
        ws.cell(row=row_idx, column=1, value=answer['student_name'])
        ws.cell(row=row_idx, column=2, value=answer['score'])
        ws.cell(row=row_idx, column=3, value=answer['total'])
        ws.cell(row=row_idx, column=4, value=percentage)
        ws.cell(row=row_idx, column=5, value=grade)
        ws.cell(row=row_idx, column=6, value=timestamp.strftime('%Y-%m-%d'))
        ws.cell(row=row_idx, column=7, value=timestamp.strftime('%I:%M %p'))
        
        # Color code based on grade
        if percentage >= 70:
            grade_color = "27AE60"
        elif percentage >= 50:
            grade_color = "F39C12"
        else:
            grade_color = "E74C3C"
        
        ws.cell(row=row_idx, column=4).font = Font(bold=True, color=grade_color)
        ws.cell(row=row_idx, column=5).font = Font(bold=True, color=grade_color)
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Return Excel file as download
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"Student_Results_{current_admin}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/logout')
def logout():
    session.pop('admin', None)
    return redirect(url_for('home'))

@app.route('/quiz')
def user_quiz():
    # Require student identification first
    if 'student_name' not in session or 'student_id' not in session:
        return redirect(url_for('start'))

    # Show list of available quizzes, but only those the student is allowed to take
    student_name = session.get('student_name')
    student_id = session.get('student_id')

    all_questions = load_questions()
    quiz_settings = load_quiz_settings()
    allowed_admins = session.get('allowed_admins', [])

    available_quizzes = []
    for admin_id, questions in all_questions.items():
        if questions and (not allowed_admins or admin_id in allowed_admins):  # Only show allowed admins (or all if not filtered)
            admin_settings = quiz_settings.get(admin_id, {'time_limit': 0})
            available_quizzes.append({
                'admin_id': admin_id,
                'admin_name': admin_id,
                'question_count': len(questions),
                'time_limit': admin_settings['time_limit']
            })

    return render_template_string(QUIZ_SELECT_TEMPLATE, available_quizzes=available_quizzes)

@app.route('/quiz/<admin_id>')
def take_quiz(admin_id):
    # Get specific admin's questions
    all_questions = load_questions()
    questions = all_questions.get(admin_id, [])
    
    # Get timer settings
    quiz_settings = load_quiz_settings()
    admin_settings = quiz_settings.get(admin_id, {'time_limit': 0})
    
    # Load existing student names for this admin (to prevent duplicates)
    all_answers = load_answers()
    admin_answers = all_answers.get(admin_id, [])
    existing_names = [entry.get('student_name','').strip() for entry in admin_answers if entry.get('student_name')]

    # Load allowed students for this admin
    allowed = load_allowed()
    allowed_list = allowed.get(admin_id, [])

    return render_template_string(USER_QUIZ_TEMPLATE, 
                                 questions=questions, 
                                 admin_name=admin_id,
                                 admin_id=admin_id,
                                 time_limit=admin_settings['time_limit'],
                                 existing_names=existing_names,
                                 allowed_list=allowed_list)

@app.route('/quiz/submit', methods=['POST'])
def submit_quiz():
    student_name = request.form.get('student_name')
    student_id = request.form.get('student_id')
    admin_id = request.form.get('admin_id')
    
    all_questions = load_questions()
    questions = all_questions.get(admin_id, [])
    
    score = 0
    results = []
    
    for i, question in enumerate(questions):
        user_answer = request.form.get(f'q{i}')
        correct = user_answer == question['correct_answer']
        if correct:
            score += 1
        
        results.append({
            'question': question['question'],
            'user_answer': user_answer,
            'correct_answer': question['correct_answer'],
            'correct': correct
        })
    
    # Save results under admin's data
    all_answers = load_answers()
    if admin_id not in all_answers:
        all_answers[admin_id] = []
    # Require student_id and verify against allowed list
    if not student_id or not student_name:
        flash('Please provide both name and student ID.', 'error')
        return redirect(url_for('take_quiz', admin_id=admin_id))

    allowed = load_allowed()
    admin_allowed = allowed.get(admin_id, [])
    name_norm = student_name.strip().lower()
    id_norm = student_id.strip().lower()
    allowed_match = any((entry.get('name') or '').strip().lower() == name_norm and (entry.get('student_id') or '').strip().lower() == id_norm for entry in admin_allowed)
    if not allowed_match:
        flash('You are not allowed to take this quiz. Please contact the instructor.', 'error')
        return redirect(url_for('take_quiz', admin_id=admin_id))

    # Enforce unique student names per admin (case-insensitive)
    normalized = name_norm
    existing = any(entry.get('student_name','').strip().lower() == normalized for entry in all_answers.get(admin_id, []))
    if existing:
        flash(f"The name '{student_name}' has already submitted this quiz.", 'error')
        return redirect(url_for('take_quiz', admin_id=admin_id))

    # Save results under admin's data
    all_answers[admin_id].append({
        'student_name': student_name,
        'score': score,
        'total': len(questions),
        'timestamp': datetime.now().isoformat(),
        'results': results
    })
    save_answers(all_answers)
    
    percentage = round((score / len(questions)) * 100, 2) if len(questions) > 0 else 0
    
    return render_template_string(RESULTS_TEMPLATE, 
                                  student_name=student_name,
                                  score=score, 
                                  total=len(questions),
                                  percentage=percentage,
                                  results=results)

if __name__ == '__main__':
    init_files()
    print("\n" + "="*50)
    print("🚀 Quiz System Starting...")
    print("="*50)
    print("\n📍 Access: http://127.0.0.1:5000")
    print("\n✨ Features:")
    print("   • Multi-admin support")
    print("   • Admin registration system")
    print("   • Quiz timer feature")
    print("   • Excel & PDF export")
    print("   • Auto-submit on timeout")
    print("\n" + "="*50 + "\n")
    app.run(debug=True, port=5000)